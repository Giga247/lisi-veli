# -*- coding: utf-8 -*-
u"""ერთჯერადი იმპორტი: xlsx + geojson -> სამი CSV Google Sheet-ისთვის.

გაშვება:  python3 tools/import.py
შედეგი:   build/plots.csv, build/users.csv, build/log.csv

სკრიპტი იდემპოტენტურია — ხელახლა გაშვება იმავე შედეგს იძლევა.
"""
import csv
import io
import json
import os
import re
import sys
import zipfile

import openpyxl

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from import_lib import (
    split_name, dedupe_by_cad, geometry_string, index_features_by_cad,
    normalize_cad, phones_from_rows)

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, u'კედრის_ქუჩა_ხელმოწერები.xlsx')
GEOJSON = os.path.join(ROOT, u'კედრის_ქუჩა_ნაკვეთები.geojson')
DOCX = os.path.join(ROOT, u'განცხადება მერიაში.docx')
OUT = os.path.join(ROOT, u'build')

def read_docx_phones(path):
    u"""ხელმოწერების დოკუმენტიდან (კოდი, ტელეფონი) წყვილები.

    სია ოთხეულებადაა აგებული: № → სახელი გვარი → საკადასტრო კოდი →
    ტელეფონი. სწორედ ამ სტრუქტურას ვეყრდნობით და არა ტექსტში ძებნას —
    დოკუმენტში თარიღიც წერია და ფართობიც, და ბრმა ძებნა მათაც აიღებდა.
    """
    if not os.path.exists(path):
        return []
    with zipfile.ZipFile(path) as archive:
        xml = archive.read('word/document.xml').decode('utf8', 'ignore')
    paragraphs = []
    for block in re.findall(r'<w:p[ >].*?</w:p>', xml, re.S):
        text = u''.join(re.findall(r'<w:t[^>]*>(.*?)</w:t>', block, re.S))
        text = re.sub(r'\s+', u' ', text).strip()
        if text:
            paragraphs.append(text)
    rows = []
    for i, text in enumerate(paragraphs):
        is_row = (re.match(r'^\d{1,3}$', text)
                  and i + 3 < len(paragraphs)
                  and re.match(r'^[\d.]{8,20}$', paragraphs[i + 2]))
        if is_row:
            rows.append((paragraphs[i + 2], paragraphs[i + 3]))
    return rows


def phone_for(cad, phone_by_cad, used):
    u"""ნაკვეთის ნომერი დოკუმენტიდან, ან ცარიელი.

    `used` მხოლოდ რეზიუმესთვისაა: ის აჩვენებს, დოკუმენტის რომელი
    ჩანაწერი ვერ დაუკავშირდა ვერცერთ ნაკვეთს — ჩუმად დაკარგული ნომერი
    მხოლოდ მაშინ აღმოჩნდებოდა, როცა ვიღაცას ვერ დაურეკავდნენ.
    """
    key = normalize_cad(cad)
    phone = phone_by_cad.get(key, u'')
    if phone:
        used.add(key)
    return phone


PLOT_HEADERS = [
    u'საკადასტრო კოდი', u'ქუჩა', u'N', u'სრული მისამართი', u'ფართობი კვ.მ',
    u'დანიშნულება', u'სახელი', u'გვარი', u'ტელეფონი', u'გრძედი', u'განედი',
    u'გეომეტრია', u'წყარო', u'შენიშვნა', u'განახლდა', u'განმაახლებელი',
]
USER_HEADERS = [
    u'მეილი', u'როლი', u'ქუჩა', u'სახელი გვარი', u'საკადასტრო კოდი',
    u'მოთხოვნის თარიღი', u'დამტკიცების თარიღი', u'დამამტკიცებელი',
]
LOG_HEADERS = [
    u'დრო', u'ვინ', u'მოქმედება', u'საკადასტრო კოდი',
    u'ველი', u'ძველი მნიშვნელობა', u'ახალი მნიშვნელობა',
]


def write_csv(path, headers, rows):
    with io.open(path, 'w', encoding='utf-8', newline='') as fh:
        writer = csv.writer(fh)
        writer.writerow(headers)
        for row in rows:
            writer.writerow(row)


def main():
    if not os.path.isdir(OUT):
        os.makedirs(OUT)

    wb = openpyxl.load_workbook(XLSX)

    plot_rows = [r for r in wb[u'ნაკვეთები'].iter_rows(min_row=2, values_only=True) if r[0]]
    sig_rows = [{u'cad': r[2], u'name': r[1]}
                for r in wb[u'ხელმოწერები'].iter_rows(min_row=2, values_only=True) if r[0]]

    sigs, dups = dedupe_by_cad(sig_rows)

    with io.open(GEOJSON, encoding='utf-8') as fh:
        geo = json.load(fh)
    geo_index, geo_dups, geo_missing = index_features_by_cad(geo[u'features'])

    phone_by_cad, phone_conflicts, phone_skipped = phones_from_rows(
        read_docx_phones(DOCX))
    phone_used = set()

    out_rows = []
    no_geometry = []
    no_area = []
    no_street = []
    for r in plot_rows:
        cad = (r[0] or u'').strip()
        sig = sigs.get(cad)
        last, first = split_name(sig[u'name'] if sig else u'')
        geom = geometry_string(geo_index.get(cad))

        if not geom:
            no_geometry.append(cad)
        if not r[4]:
            no_area.append(cad)
        if not r[1]:
            no_street.append(cad)

        out_rows.append([
            cad,            # საკადასტრო კოდი
            r[1] or u'',    # ქუჩა
            r[2] or u'',    # N
            r[3] or u'',    # სრული მისამართი
            r[4] or u'',    # ფართობი კვ.მ
            r[5] or u'',    # დანიშნულება
            first,          # სახელი
            last,           # გვარი
            phone_for(cad, phone_by_cad, phone_used),  # ტელეფონი
            r[9] or u'',    # გრძედი
            r[10] or u'',   # განედი
            geom,           # გეომეტრია
            r[7] or u'',    # წყარო
            r[8] or u'',    # შენიშვნა
            u'',            # განახლდა
            u'',            # განმაახლებელი
        ])

    write_csv(os.path.join(OUT, u'plots.csv'), PLOT_HEADERS, out_rows)
    write_csv(os.path.join(OUT, u'users.csv'), USER_HEADERS, [])
    write_csv(os.path.join(OUT, u'log.csv'), LOG_HEADERS, [])

    with_owner = sum(1 for r in out_rows if r[7])
    print(u'--- იმპორტის რეზიუმე ---')
    print(u'ნაკვეთი:            %d' % len(out_rows))
    print(u'მფლობელით:          %d' % with_owner)
    print(u'პოლიგონით:          %d' % (len(out_rows) - len(no_geometry)))
    print(u'პოლიგონის გარეშე:   %d  %s' % (len(no_geometry), no_geometry))
    print(u'ფართობის გარეშე:    %d  %s' % (len(no_area), no_area))
    print(u'ქუჩის გარეშე:       %d  %s' % (len(no_street), no_street))
    print(u'დუბლიკატი კოდი:     %d  %s' % (len(dups), dups))
    print(u'ტელეფონით:          %d' % len(phone_used))
    print(u'ტელეფონის გარეშე:   %d' % (len(out_rows) - len(phone_used)))
    unused = sorted(set(phone_by_cad) - phone_used)
    if unused:
        print(u'დოკუმენტში ნომერი აქვს, ბაზაში ნაკვეთი არ არის: %d  %s'
              % (len(unused), unused))
    if phone_skipped:
        print(u'ნომერი ვერ წაიკითხა (ტიპო ან ცარიელი): %d  %s'
              % (len(phone_skipped), phone_skipped))
    if phone_conflicts:
        print(u'ერთ ნაკვეთზე ორი სხვადასხვა ნომერი: %d  %s'
              % (len(phone_conflicts), phone_conflicts))
    print(u'დუბლიკატი geojson-ში: %d  %s' % (len(geo_dups), geo_dups))
    print(u'geojson feature კოდის გარეშე: %d' % geo_missing)
    print(u'')
    print(u'ხელით გადასამოწმებელი: სამსიტყვიანი სახელები')
    for cad, sig in sigs.items():
        if len((sig[u'name'] or u'').split()) > 2:
            print(u'   %s  %s' % (cad, sig[u'name']))
    print(u'')
    print(u'ჩაწერილია: %s' % OUT)


if __name__ == '__main__':
    main()
